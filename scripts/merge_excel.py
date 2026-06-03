#!/usr/bin/env python3
"""
Incremental merge: scans output/ for per-video Excel files, appends only BVs
not already in naafiri_mid_index_merged.xlsx. Existing rows/images are preserved.
"""
import os, sys, glob, re, io
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

OUTPUT_DIR = "output"
MERGED_NAME = "naafiri_mid_index_merged.xlsx"
SHEET_NAME = "Naafiri中单索引"

COL_HEADERS = ["序号", "视频标题", "BV号", "跳转链接", "对局时间(分P内)", "分P", "对位英雄", "截图"]
COL_WIDTHS  = [4,      38,       16,     50,       12,               4,    18,        22]
ROW_HEIGHT  = 100

# Source column map
SRC_TITLE    = 2
SRC_BVID     = 3
SRC_LINK     = 4
SRC_TIME     = 5
SRC_PART     = 6
SRC_OPPONENT = 8
SRC_IMAGE    = 14

# ─── Styles ────────────────────────────────────────────────────────────
HDR_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(size=9)
CELL_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LINK_FONT = Font(size=9, color="0000FF", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_cell(cell, is_header=False):
    cell.font = HDR_FONT if is_header else CELL_FONT
    if is_header:
        cell.fill = HDR_FILL
    cell.border = THIN_BORDER
    cell.alignment = CELL_ALIGN


def cn_only(text):
    """Extract Chinese name from 'Galio / 正义巨像' → '正义巨像'.
       If no '/ ' separator, return as-is."""
    s = str(text or "")
    if " / " in s:
        return s.split(" / ", 1)[1]
    return s


# ─── Read existing merged file ─────────────────────────────────────────
def read_existing_merged(path):
    """Read existing merged file. Returns (existing_bvs, existing_row_count)."""
    if not os.path.isfile(path):
        return set(), 0
    wb = load_workbook(path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    bvs = set()
    row_count = ws.max_row - 1  # minus header
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, 3).value  # BV号 is col 3
        if val:
            bvs.add(str(val).strip())
    wb.close()
    return bvs, max(row_count, 0)


# ─── Read source Excel ─────────────────────────────────────────────────
def read_source_sheet(filepath):
    """Read one per-video Excel. Returns (rows, images)."""
    wb = load_workbook(filepath)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {}
        row["title"]    = ws.cell(r, SRC_TITLE).value
        row["bvid"]     = ws.cell(r, SRC_BVID).value
        row["time"]     = ws.cell(r, SRC_TIME).value
        row["part"]     = ws.cell(r, SRC_PART).value
        row["opponent"] = cn_only(ws.cell(r, SRC_OPPONENT).value)
        # Hyperlink
        hl = ws.cell(r, SRC_LINK).hyperlink
        row["url"] = hl.target if hl else str(ws.cell(r, SRC_LINK).value or "")
        if row["bvid"] is None and row["title"] is None:
            continue
        rows.append(row)

    # Images
    images = []
    for img in ws._images:
        anchor_row = img.anchor._from.row + 1
        img_bytes = io.BytesIO(img._data())
        images.append({
            "src_row": anchor_row,
            "bytes": img_bytes,
            "width": img.width,
            "height": img.height,
        })
    wb.close()
    return rows, images


# ─── Main ──────────────────────────────────────────────────────────────
def main():
    # 1. Find source files
    pattern = os.path.join(OUTPUT_DIR, "naafiri_mid_index_BV*_final_v2.xlsx")
    src_files = sorted(glob.glob(pattern))
    if not src_files:
        print("[ERROR] No source Excel files in output/")
        sys.exit(1)
    print(f"Source files: {len(src_files)}")

    # 2. Read existing merged state
    merged_path = MERGED_NAME
    existing_bvs, existing_count = read_existing_merged(merged_path)
    if existing_bvs:
        print(f"Existing merged: {existing_count} rows, {len(existing_bvs)} BVs")
    else:
        print("No existing merged file — creating new.")

    # 3. Determine which BVs are new
    all_source_bvs = {}
    for f in src_files:
        bv_match = re.search(r'(BV[a-zA-Z0-9]+)', os.path.basename(f))
        if bv_match:
            all_source_bvs[bv_match.group(1)] = f

    new_bvs = {bv: f for bv, f in all_source_bvs.items() if bv not in existing_bvs}
    skipped = [bv for bv in all_source_bvs if bv in existing_bvs]

    if skipped:
        print(f"Already merged (skip): {len(skipped)} — {', '.join(skipped)}")
    if not new_bvs:
        print("No new BVs to merge. Done.")
        return

    print(f"New BVs to merge: {len(new_bvs)} — {', '.join(new_bvs.keys())}")

    # 4. Read new source files
    new_rows = []
    new_images = []
    per_bv_counts = {}
    base_seq = existing_count

    for bv, f in sorted(new_bvs.items()):
        rows, images = read_source_sheet(f)
        offset = len(new_rows)  # relative to new_rows
        for row in rows:
            row["_bv"] = bv
        new_rows.extend(rows)
        # Shift image row targets
        for img in images:
            src_row = img["src_row"]
            target_row = base_seq + offset + (src_row - 2) + 2
            new_images.append({
                "row": target_row,
                "bytes": img["bytes"],
                "width": img["width"],
                "height": img["height"],
            })
        per_bv_counts[bv] = len(rows)
        print(f"  {bv}: {len(rows)} rows, {len(images)} images")

    total_new = len(new_rows)
    print(f"Total new: {total_new} rows, {len(new_images)} images")

    # 5. Open or create merged workbook
    if os.path.isfile(merged_path):
        wb = load_workbook(merged_path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        # Write headers
        for c, h in enumerate(COL_HEADERS, 1):
            cell = ws.cell(1, c, h)
            style_cell(cell, is_header=True)
        for c, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

    # 6. Append new rows
    start_row = existing_count + 2  # 1-indexed, after header & existing rows

    for i, row in enumerate(new_rows):
        r = start_row + i
        seq = existing_count + i + 1

        cell = ws.cell(r, 1, seq);           style_cell(cell)                   # 序号
        cell = ws.cell(r, 2, row["title"]);  style_cell(cell); cell.alignment = CELL_ALIGN_LEFT  # 标题
        cell = ws.cell(r, 3, row["bvid"]);   style_cell(cell)                   # BV号
        cell = ws.cell(r, 4, row["url"]);    style_cell(cell)                   # 跳转
        if row["url"]:
            cell.hyperlink = row["url"]; cell.font = LINK_FONT
        cell = ws.cell(r, 5, row["time"]);   style_cell(cell)                   # 时间
        cell = ws.cell(r, 6, row["part"]);   style_cell(cell)                   # 分P
        cell = ws.cell(r, 7, row["opponent"]); style_cell(cell)                 # 对位(中文)

    # 7. Append new images
    for img_info in new_images:
        target_row = img_info["row"]
        try:
            img = XLImage(img_info["bytes"])
            scale = 0.18
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
            img.anchor = f"H{target_row}"
            ws.add_image(img)
        except Exception as e:
            print(f"  [WARN] Image at row {target_row}: {e}")
        ws.row_dimensions[target_row].height = ROW_HEIGHT

    # 8. Update freeze & filter to cover all rows
    total_rows = existing_count + total_new
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{total_rows + 1}"

    # 9. Save
    wb.save(merged_path)
    print(f"\nSaved: {merged_path}")
    print(f"  Before: {existing_count} rows")
    print(f"  Added:  {total_new} rows ({len(new_bvs)} videos)")
    print(f"  Total:  {total_rows} rows, {len(new_images)} new screenshots")


if __name__ == "__main__":
    main()
